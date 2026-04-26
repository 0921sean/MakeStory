"""
KorA Valley 계획/완료 트래킹 시스템

카카오톡 채팅 로그(tracking.txt)를 파싱하여 멤버들의 계획 작성 및 완료(✅)를
엑셀 DB + 계획 시트 + 개인 시트에 자동 반영합니다.

주요 기능:
    - 계획 작성 → DB 시트에 행 추가, 계획 시트에 날짜 기록, 개인 시트에 내용 기록
    - 완료 처리(✅) → DB 상태 변경, 계획 시트에 O 표시, 개인 시트에 취소선
    - 주간 리포트 (TOP 3 완료왕, 무활동자)
    - tracking.txt 자동 백업(archive)
"""

import os
import re
import shutil
import sys
from collections import Counter
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

# 공통 모듈 import
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from common.config import KORA_NAME_MAP

# =============================
# 경로 설정
# =============================
excel_path = "./../KorA_Valley/KorA_Valley_tracking_2026_04_19_DB.xlsx"
output_path = "./../KorA_Valley/KorA_Valley_tracking_2026_04_26_DB.xlsx"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TRACKING_FILE = os.path.join(BASE_DIR, "tracking.txt")
ARCHIVE_DIR = os.path.join(BASE_DIR, "plan_archive")

# =============================
# 멤버 매핑 (config에서 가져옴)
# =============================
name_to_sheet = KORA_NAME_MAP
RAW_NAMES = list(name_to_sheet.keys())
SHEET_NAMES = list(name_to_sheet.values())

# =============================
# 엑셀 로드
# =============================
wb = load_workbook(excel_path)
db_sheet = wb["DB"]
plan_sheet = wb["계획"]

# =============================
# 스타일
# =============================
BLACK_FONT = Font(color="FF000000")
RED_CENTER = Font(color="FFFF0000")
CENTER = Alignment(horizontal="center")
RIGHT = Alignment(horizontal="right")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# =============================
# ⭐ 신규 멤버(시트) 자동 추가
# =============================
for i, name in enumerate(SHEET_NAMES):
    col = i + 3
    
    # 1. 계획 시트에 해당 멤버 열(이름)이 없는 경우 세팅
    if not plan_sheet.cell(2, col).value:
        date_str = input(f"\n새 멤버 '{name}'의 들어온 날짜를 입력하세요 (예: 2026.4.26): ")
        
        # 날짜 (1행)
        date_cell = plan_sheet.cell(1, col, date_str)
        date_cell.alignment = CENTER
        date_cell.border = THIN_BORDER
        
        # 이름 (2행)
        name_cell = plan_sheet.cell(2, col, name)
        name_cell.alignment = CENTER
        name_cell.border = THIN_BORDER

    # 2. 개인 시트가 없는 경우 생성 및 세팅
    if name not in wb.sheetnames:
        print(f"[{name}] 개인 시트를 맨 뒤에 생성하고 초기화합니다.")
        new_sheet = wb.create_sheet(title=name)
        
        # B2: 날짜, C2: 이름
        date_cell2 = new_sheet.cell(2, 2, plan_sheet.cell(1, col).value)
        date_cell2.alignment = CENTER
        date_cell2.border = THIN_BORDER
        
        name_cell2 = new_sheet.cell(2, 3, name)
        name_cell2.alignment = CENTER
        name_cell2.border = THIN_BORDER

person_sheets = {
    name: wb[name]
    for name in SHEET_NAMES
    if name in wb.sheetnames
}



# =============================
# ⭐ STEP 0: 계획 시트 기존 빨간색 → 검은색 리셋
# =============================
print("🔄 계획 시트: 기존 빨간색 → 검은색 리셋 중...")
reset_count = 0

for col_idx in range(3, 3 + len(SHEET_NAMES)):
    for row_idx in range(3, plan_sheet.max_row + 1):
        cell = plan_sheet.cell(row_idx, col_idx)
        if cell.value is not None and cell.font and cell.font.color:
            color = cell.font.color
            if color.rgb and "FF0000" in str(color.rgb):
                cell.font = Font(color="FF000000")
                cell.alignment = CENTER
                reset_count += 1

print(f"   → {reset_count}개 셀 리셋 완료")


# =============================
# 날짜 기록 함수
# =============================
def write_db_date(sheet, r, c, dt):
    sheet.cell(r, c, dt).number_format = "yyyy.mm.dd"


def write_plan_date(sheet, r, c, dt):
    cell = sheet.cell(r, c, dt)
    cell.number_format = 'm"월" d"일"'
    cell.font = RED_CENTER
    cell.alignment = CENTER


# =============================
# DB row 찾기
# =============================
def find_db_row(name, plan_no):
    result = None
    for r in range(2, db_sheet.max_row + 1):
        if db_sheet.cell(r, 2).value == name and db_sheet.cell(r, 3).value == plan_no:
            result = r  # 매칭될 때마다 갱신 → 루프 끝나면 마지막 row
    return result


# =============================
# tracking.txt 백업
# =============================
date_match = re.search(r"(\d{4}_\d{2}_\d{2})", output_path)
if not date_match:
    raise ValueError("output_path에서 날짜 추출 실패")

os.makedirs(ARCHIVE_DIR, exist_ok=True)
archive_path = os.path.join(ARCHIVE_DIR, f"archive_{date_match.group(1)}.txt")
shutil.copy(TRACKING_FILE, archive_path)
print(f"📦 archive 저장 완료 → {archive_path}")

# =============================
# TXT 파싱
# =============================
with open(TRACKING_FILE, encoding="utf-8") as f:
    lines = f.read().splitlines()

current_date = None

# 이번 주 활동 추적용
weekly_completions = Counter()
weekly_active = set()
unrecognized_lines = []

for original_line in lines:
    line = re.sub(r"\s+", " ", original_line.strip())
    if not line:
        continue

    # 날짜
    m = re.match(r"(\d{4}년 \d{1,2}월 \d{1,2}일)", line)
    if m:
        current_date = datetime.strptime(m.group(1), "%Y년 %m월 %d일")
        continue

    # 시간 제거
    line = re.sub(r"^(오전|오후) \d{1,2}:\d{2} ", "", line)

    raw_name, name, content = None, None, None
    for candidate in RAW_NAMES:
        if line.startswith(candidate + " "):
            raw_name = candidate
            name = name_to_sheet[candidate]
            content = line[len(candidate) + 1 :].strip()
            break

    if not name:
        unrecognized_lines.append(original_line)
        continue

    is_recognized = False

    # -------------------------
    # 계획 작성
    # -------------------------
    plan_match = re.search(r"(\d+)\s*번[째쨰]?\s*(계획|게획|목표)\s*:?\s*(.*)", content)
    if plan_match and "✅" not in content:
        is_recognized = True
        plan_no = int(plan_match.group(1))
        plan_text = plan_match.group(3).strip()

        weekly_active.add(name)

        # DB 시트: 내용이 다르면 새 행 추가
        existing_row = find_db_row(name, plan_no)
        if not existing_row or db_sheet.cell(existing_row, 4).value != plan_text:
            r = db_sheet.max_row + 1
            db_sheet.append([r - 1, name, plan_no, plan_text, "미완료", None, None])
            write_db_date(db_sheet, r, 6, current_date)

        # 계획 시트 날짜
        col = SHEET_NAMES.index(name) + 3
        row = plan_no + 2
        if not plan_sheet.cell(row, col).value:
            write_plan_date(plan_sheet, row, col, current_date)

        # 개인 시트 계획 내용
        if name in person_sheets:
            person_sheets[name].cell(row, 3, plan_text)

    # -------------------------
    # 완료 처리
    # -------------------------
    if "✅" in content:
        m = re.search(r"(\d+)\s*번", content)
        if m:
            is_recognized = True
            plan_no = int(m.group(1))

            weekly_active.add(name)
            weekly_completions[name] += 1

            r = find_db_row(name, plan_no)
            if r:
                db_sheet.cell(r, 5, "완료")
                write_db_date(db_sheet, r, 7, current_date)

            col = SHEET_NAMES.index(name) + 3
            row = plan_no + 2
            plan_sheet.cell(row, col, "O").font = RED_CENTER
            plan_sheet.cell(row, col).alignment = CENTER

            if name in person_sheets:
                c = person_sheets[name].cell(row, 3)
                if c.value:
                    c.font = Font(strike=True)

    if not is_recognized:
        unrecognized_lines.append(original_line)

# =============================
# 계획 시트 번호 확장 (B열만)
# =============================
max_plan_no = max(
    [db_sheet.cell(r, 3).value for r in range(2, db_sheet.max_row + 1) if isinstance(db_sheet.cell(r, 3).value, int)],
    default=0,
)

for n in range(1, max_plan_no + 1):
    r = n + 2
    b = plan_sheet.cell(r, 2, n)
    b.alignment = RIGHT
    b.border = THIN_BORDER

# =============================
# 각자 시트 번호 확장 (개인별)
# =============================
for sheet in person_sheets.values():
    last_row = max(
        [r for r in range(3, sheet.max_row + 1) if sheet.cell(r, 3).value],
        default=2,
    )
    max_no = last_row - 2
    for n in range(1, max_no + 1):
        r = n + 2
        b = sheet.cell(r, 2, n)
        b.alignment = RIGHT
        b.border = THIN_BORDER

# =============================
# 저장
# =============================
wb.save(output_path)
print("✅ 전체 시트(DB + 계획 + 개인 시트) 업데이트 완료")

# =============================
# 주간 리포트 출력
# =============================
print("\n" + "=" * 50)
print("📊 이번 주 정산 리포트")
print("=" * 50)

# TOP 3 완료왕
print("\n🏆 이번 주 완료 TOP 3:")
if weekly_completions:
    ranked = weekly_completions.most_common()

    prev_count = None
    rank = 0
    medals = {1: "🥇", 2: "🥈", 3: "🥉"}

    for i, (person, count) in enumerate(ranked):
        if count != prev_count:
            rank = i + 1
        if rank > 3:
            break
        medal = medals.get(rank, "  ")
        print(f"   {medal} {rank}등: {person} ({count}개 완료)")
        prev_count = count
else:
    print("   (이번 주 완료 없음)")

# 무활동자
print("\n😴 이번 주 활동 없는 멤버:")
inactive = [name for name in SHEET_NAMES if name not in weekly_active]
if inactive:
    for person in inactive:
        print(f"   ⚠️  {person}")
    print(f"\n   → 총 {len(inactive)}명 / {len(SHEET_NAMES)}명 무활동")
else:
    print("   🎉 전원 활동! 모두 참여했습니다!")

# 인식 안 된 줄 출력
if unrecognized_lines:
    print("\n" + "=" * 50)
    print("⚠️  인식되지 않은 줄 목록 (포맷 불일치, 일반 대화 등)")
    print("=" * 50)
    for u_line in unrecognized_lines:
        print(u_line)

print("\n" + "=" * 50)