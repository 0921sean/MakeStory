import openpyxl
wb = openpyxl.load_workbook('./../KorA_Valley/KorA_Valley_tracking_2026_04_19_DB.xlsx')
plan = wb['계획']
print("계획 시트 C1:C2")
print(plan.cell(1, 3).value, plan.cell(2, 3).value)

if '천승범' in wb.sheetnames:
    sheet = wb['천승범']
    print("천승범 시트 B2, C2")
    print(sheet.cell(2, 2).value, sheet.cell(2, 3).value)
