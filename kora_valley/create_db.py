"""
KorA Valley DB 시트 초기 생성

기존 엑셀의 계획 시트 + 개인 시트 데이터를 기반으로
DB 시트를 새로 생성합니다.
"""

import os
import sys
from datetime import datetime

from openpyxl import load_workbook

# 공통 모듈 import
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from common.config import KORA_NAME_MAP

excel_path = "./../KorA_Valley/KorA_Valley_tracking_2026_02_08.xlsx"
output_path = "./../KorA_Valley/KorA_Valley_tracking_2026_02_08_with_DB.xlsx"

names = list(KORA_NAME_MAP.values())

wb = load_workbook(excel_path)

if "DB" in wb.sheetnames:
    del wb["DB"]

db_sheet = wb.create_sheet("DB")

headers = ["ID", "이름", "계획번호", "계획내용", "상태", "생성일", "완료일"]
for col, header in enumerate(headers, start=1):
    db_sheet.cell(row=1, column=col, value=header)

plan_sheet = wb["계획"]

current_id = 1
db_row = 2


def write_standard_date(sheet, row, col, value):
    """다양한 형식의 날짜 문자열을 yyyy.mm.dd 형식으로 통일하여 기록합니다."""
    if not value:
        sheet.cell(row=row, column=col, value=None)
        return

    # 이미 datetime이면 그대로 사용
    if isinstance(value, datetime):
        dt = value

    # "5월 1일" 형태면 변환
    elif isinstance(value, str) and "월" in value:
        try:
            parts = value.replace("월", "").replace("일", "").split()
            month = int(parts[0])
            day = int(parts[1])
            year = datetime.now().year
            dt = datetime(year, month, day)
        except Exception:
            sheet.cell(row=row, column=col, value=value)
            return

    # "2025.4.14" 형태면 변환
    elif isinstance(value, str) and "." in value:
        try:
            dt = datetime.strptime(value, "%Y.%m.%d")
        except ValueError:
            try:
                dt = datetime.strptime(value, "%Y.%m.%d.")
            except ValueError:
                sheet.cell(row=row, column=col, value=value)
                return
    else:
        sheet.cell(row=row, column=col, value=value)
        return

    sheet.cell(row=row, column=col, value=dt)
    sheet.cell(row=row, column=col).number_format = "yyyy.mm.dd"


# -------------------------
# DB 생성
# -------------------------
for name_idx, name in enumerate(names):
    person_sheet_name = name
    if person_sheet_name not in wb.sheetnames:
        continue

    person_sheet = wb[person_sheet_name]
    plan_col = name_idx + 3

    for row in range(3, plan_sheet.max_row + 1):
        plan_number = plan_sheet.cell(row=row, column=2).value
        status_cell = plan_sheet.cell(row=row, column=plan_col).value

        if not status_cell:
            continue

        # 개인 시트에서 계획내용 찾기
        plan_content = None
        for r in range(3, person_sheet.max_row + 1):
            if person_sheet.cell(row=r, column=2).value == plan_number:
                plan_content = person_sheet.cell(row=r, column=3).value
                break

        # 상태 판별
        if str(status_cell).strip().upper() == "O":
            status = "완료"
            생성일 = None
            완료일 = None
        else:
            status = "미완료"
            생성일 = status_cell
            완료일 = None

        db_sheet.cell(row=db_row, column=1, value=current_id)
        db_sheet.cell(row=db_row, column=2, value=name)
        db_sheet.cell(row=db_row, column=3, value=plan_number)
        db_sheet.cell(row=db_row, column=4, value=plan_content)
        db_sheet.cell(row=db_row, column=5, value=status)

        write_standard_date(db_sheet, db_row, 6, 생성일)
        write_standard_date(db_sheet, db_row, 7, 완료일)

        current_id += 1
        db_row += 1

wb.save(output_path)
print("✅ DB 생성 완료 (정확 완료 판별 적용)")