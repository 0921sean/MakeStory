import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import os
import shutil

excel_path = './../KorA_Valley/KorA_Valley_tracking_2026_02_08_DB_create.xlsx'
output_path = './../KorA_Valley/KorA_Valley_tracking_2026_02_13_DB.xlsx'

name_to_sheet = {
    "천승범": "천승범",
    "비씩 20 조예찬형": "조예찬",
    "신재욱": "신재욱",
    "서희찬": "서희찬",
    "윤상민": "윤상민",
    "최서연": "최서연",
    "이연희 Kirsten": "이연희",
    "김영준": "김영준",
    "정성민": "정성민",
    "김진호": "김진호",
    "유재호님": "유재호",
    "황서호형": "황서호",
    "박상준님": "박상준",
    "홍석채님": "홍석채",
    "전상우님": "전상우",
    "김세현님": "김세현",
    "비씩 20 배경덕": "배경덕"
}

names = list(name_to_sheet.keys())

wb = load_workbook(excel_path)
db_sheet = wb["DB"]
plan_sheet = wb["계획"]

# -------------------------
# 날짜 기록 함수
# -------------------------

def write_db_date(sheet, row, col, dt):
    sheet.cell(row=row, column=col, value=dt)
    sheet.cell(row=row, column=col).number_format = "yyyy.mm.dd"

def write_plan_date(sheet, row, col, dt):
    cell = sheet.cell(row=row, column=col, value=dt)
    cell.number_format = 'm"월" d"일"'
    cell.font = Font(color="FFFF0000")
    cell.alignment = Alignment(horizontal="center")

# -------------------------
# DB 행 찾기
# -------------------------

def find_db_row(name, plan_number):
    for r in range(2, db_sheet.max_row + 1):
        if (str(db_sheet.cell(r,2).value).strip() == str(name).strip() and
            str(db_sheet.cell(r,3).value).strip() == str(plan_number).strip()):
            return r
    return None

# ----------------------------
# output_path에서 날짜 추출
# ----------------------------

file_name = os.path.basename(output_path)

# 2026_02_11 패턴 찾기
date_match = re.search(r'(\d{4}_\d{2}_\d{2})', file_name)

if not date_match:
    raise ValueError("output_path에서 날짜를 찾을 수 없습니다.")

date_str = date_match.group(1)

# ----------------------------
# plan_archive 경로 설정
# ----------------------------

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
archive_folder = os.path.join(BASE_DIR, "plan_archive")

if not os.path.exists(archive_folder):
    os.makedirs(archive_folder)

archive_path = os.path.join(
    archive_folder,
    f"archive_{date_str}.txt"
)

# 원본 파일
source_file = os.path.join(BASE_DIR, "tracking.txt")

shutil.copy(source_file, archive_path)

print(f"📦 archive 저장 완료 → {archive_path}")

# -------------------------
# TXT 읽기
# -------------------------

with open('kora_valley/tracking.txt', 'r', encoding='utf-8') as f:
    lines = f.read().strip().split('\n')

current_date = None

for line in lines:

    line = re.sub(r'\s+', ' ', line.strip())
    if not line:
        continue

    # 날짜 줄
    date_match = re.match(r'(\d{4}년 \d{1,2}월 \d{1,2}일)', line)
    if date_match:
        current_date = datetime.strptime(date_match.group(1), "%Y년 %m월 %d일")
        continue

    # 시간 제거
    line = re.sub(r'^(오전|오후) \d{1,2}:\d{2} ', '', line)

    # 이름 찾기
    name = None
    content = None

    for candidate in names:
        if line.startswith(candidate + ' '):
            name = name_to_sheet[candidate]  # value 값만 사용
            content = line[len(candidate)+1:].strip()
            break

    if not name:
        continue

    # -------------------------
    # 계획 작성
    # -------------------------

    plan_match = re.search(r'(\d+)\s*번[째쨰]?\s*(계획|목표)\s*:?\s*(.*)', content)

    if plan_match and '✅' not in content:

        plan_number = int(plan_match.group(1))
        plan_content = plan_match.group(3).strip()

        if not find_db_row(name, plan_number):

            new_row = db_sheet.max_row + 1

            db_sheet.cell(new_row,1,new_row-1)  # ID
            db_sheet.cell(new_row,2,name)
            db_sheet.cell(new_row,3,plan_number)
            db_sheet.cell(new_row,4,plan_content)
            db_sheet.cell(new_row,5,"미완료")

            write_db_date(db_sheet,new_row,6,current_date)
            db_sheet.cell(new_row,7,None)

        # 계획 시트 반영
        name_index = list(name_to_sheet.values()).index(name)
        col = name_index + 3
        row_idx = plan_number + 2

        if not plan_sheet.cell(row_idx,col).value:
            write_plan_date(plan_sheet,row_idx,col,current_date)

    # -------------------------
    # 완료 처리 (✅ 기준)
    # -------------------------

    if '✅' in content:

        completion_match = re.search(r'(\d+)\s*번', content)

        if completion_match:
            plan_number = int(completion_match.group(1))
            db_row = find_db_row(name, plan_number)

            if db_row:
                db_sheet.cell(db_row,5,"완료")
                write_db_date(db_sheet,db_row,7,current_date)

            # 계획 시트 O 표시
            name_index = list(name_to_sheet.values()).index(name)
            col = name_index + 3
            row_idx = plan_number + 2

            cell = plan_sheet.cell(row_idx,col,"O")
            cell.font = Font(color="FFFF0000")
            cell.alignment = Alignment(horizontal="center")
            
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ----------------------------
# DB 기준 최대 계획번호 찾기
# ----------------------------
max_plan_number = 0
for r in range(2, db_sheet.max_row + 1):
    val = db_sheet.cell(r,3).value
    if isinstance(val, int):
        max_plan_number = max(max_plan_number, val)

print("최대 계획번호:", max_plan_number)

# ----------------------------
# 기존 마지막 번호 찾기
# ----------------------------
current_max_number = 0
for r in range(3, plan_sheet.max_row + 1):
    val = plan_sheet.cell(r,2).value
    if isinstance(val, int):
        current_max_number = max(current_max_number, val)

# ----------------------------
# 부족한 번호만 추가 (B열만 테두리)
# ----------------------------
if max_plan_number > current_max_number:
    for num in range(current_max_number + 1, max_plan_number + 1):
        row_idx = num + 2  # 계획번호 1 -> row 3

        # B열 번호
        b = plan_sheet.cell(row=row_idx, column=2, value=num)
        b.alignment = Alignment(horizontal="right")

        # ✅ B열에만 테두리 적용 (행 전체 X)
        b.border = thin_border

wb.save(output_path)
print("✅ DB + 계획 시트 동시 업데이트 완료")
