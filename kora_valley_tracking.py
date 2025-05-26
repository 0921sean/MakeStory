import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from datetime import datetime
from openpyxl.styles import Alignment

# 파일 경로
excel_path = './../KorA_Valley/KorA_Valley_tracking.xlsx'  # 수정
output_path = './../KorA_Valley/KorA_Valley_tracking_fixed.xlsx'

# 사람 이름 목록
names = ["천승범", "조예찬 John", "최재훈", "송의현", "이승준", "양수민", "신재욱", "이승헌", "진세", "서희찬", "신영진", "윤상민"]

# 사람 이름과 시트 이름 매핑 (직접 설정)
name_to_sheet = {
    "천승범": "천승범",
    "조예찬 John": "조예찬",
    "최재훈": "최재훈",
    "송의현": "송의현",
    "이승준": "이승준",
    "양수민": "양수민",
    "신재욱": "신재욱",
    "이승헌": "이승헌",
    "진세": "김진세",  # '진세'는 '김진세' 시트에 기록
    "서희찬": "서희찬",
    "신영진": "신영진",
    "윤상민": "윤상민"
}

# 엑셀 파일 읽기
wb = load_workbook(excel_path)

# '계획' 시트 찾기
if '계획' in wb.sheetnames:
    plan_sheet = wb['계획']
else:
    print("'계획' 시트를 찾을 수 없습니다. 첫 번째 시트를 사용합니다.")
    plan_sheet = wb.active

# 각 사람별 시트 찾기 및 매핑
name_sheets = {}
for name, sheet_name in name_to_sheet.items():
    if sheet_name in wb.sheetnames:
        name_sheets[name] = wb[sheet_name]
    else:
        print(f"'{sheet_name}' 시트를 찾을 수 없습니다. 해당 사람의 계획 내용은 기록되지 않습니다: {name}")

# 입력 텍스트 읽기
with open('tracking.txt', 'r', encoding='utf-8') as f:
    tracking_list = f.read()

current_date = None
lines = tracking_list.strip().split('\n')

for line in lines:
    line = re.sub(r'\s+', ' ', line.strip())  # 공백 여러개 정리
    if not line:
        continue

    # 날짜 줄
    date_match = re.match(r'(\d{4}년 \d{1,2}월 \d{1,2}일)', line)
    if date_match:
        date_obj = datetime.strptime(date_match.group(1), "%Y년 %m월 %d일")
        current_date = f"{date_obj.month}월 {date_obj.day}일"
        continue
    
    # 이름 자동 매칭
    name = None
    content = None

    # 여기서 시간(오전/오후 시간) 제거
    line = re.sub(r'^(오전|오후) \d{1,2}:\d{2} ', '', line)

    for candidate in names:
        if line.startswith(candidate + ' '):  # 이름 뒤에 공백 확인
            name = candidate
            content = line[len(candidate)+1:].strip()  # 이름+공백만큼 잘라서 content 추출
            break

    if not name or not content:
        continue  # 이름 못 찾았으면 건너뜀

    name_idx = names.index(name)
    col = name_idx + 3  # 열 인덱스 (3열부터 시작)

    # 계획 패턴 확인 - "번째 계획:" 형식만 처리
    plan_match = re.search(r'(\d+)번[째쨰]?\s*(계획|목표)\s*:?\s*(.*)', content)
    if plan_match and '✅' not in content:  # 완료 표시가 없는 경우만 계획으로 처리
        plan_number = int(plan_match.group(1))
        plan_content = plan_match.group(3).strip()
        row = plan_number + 2  # 행 인덱스 (3행부터 시작)
        
        # '계획' 시트에 날짜 기록
        if not plan_sheet.cell(row=row, column=col).value:
            cell = plan_sheet.cell(row=row, column=col, value=current_date)
            cell.font = Font(color="FFFF0000")
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = '@'
        
        # 해당 이름의 시트가 있다면 계획 내용 기록 - 3번째 열에 기록
        if name in name_sheets and plan_content:  # 계획 내용이 있는 경우에만 기록
            cell = name_sheets[name].cell(row=row, column=3, value=plan_content)  # 3번째 열로 수정
            # 계획 페이지를 제외한 다른 페이지는 빨간색 글씨 사용 안 함
            cell.alignment = Alignment(horizontal="left")

    # 완료 처리
    if '✅' in content:
        completion_match = re.search(r'✅\s*(\d+)번[째쨰]?\s*(계획|목표)\s*완료', content)
        if completion_match:
            plan_number = int(completion_match.group(1))
            row = plan_number + 2
            
            # '계획' 시트에 완료 표시
            cell = plan_sheet.cell(row=row, column=col, value='O')
            cell.font = Font(color="FFFF0000")
            cell.alignment = Alignment(horizontal="center")

# 모든 시트에서 가장 높은 계획 번호 찾기
max_plan_number = 0
all_sheets = [plan_sheet] + list(name_sheets.values())

for sheet in all_sheets:
    # 각 시트에서 내용이 있는 마지막 행 찾기
    for row in range(sheet.max_row, 2, -1):  # 마지막 행부터 3행까지 역순으로 검사
        has_content = False
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value:
                has_content = True
                break
        if has_content:
            # 행 번호에서 계획 번호 계산 (행번호 - 2)
            plan_number = row - 2
            max_plan_number = max(max_plan_number, plan_number)
            break

print(f"가장 높은 계획 번호: {max_plan_number}")

# 모든 시트의 2번째 열에 번호 추가 및 테두리 적용
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

all_sheets = [plan_sheet] + list(name_sheets.values())

# 먼저 모든 시트에서 가장 높은 계획 번호 찾기 ('계획' 시트용)
global_max_plan_number = 0
for sheet in all_sheets:
    # 각 시트에서 내용이 있는 마지막 행 찾기
    for row in range(sheet.max_row, 2, -1):  # 마지막 행부터 3행까지 역순으로 검사
        has_content = False
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value:
                has_content = True
                break
        if has_content:
            # 행 번호에서 계획 번호 계산 (행번호 - 2)
            plan_number = row - 2
            global_max_plan_number = max(global_max_plan_number, plan_number)
            break

for sheet in all_sheets:
    # '계획' 시트인지 확인
    is_plan_sheet = (sheet == plan_sheet)
    
    if is_plan_sheet:
        # '계획' 시트는 전체 최대 계획 번호 사용
        max_plan_number_in_sheet = global_max_plan_number
    else:
        # 다른 시트들은 해당 시트에서 3번째 열의 마지막 내용이 있는 행 찾기
        max_content_row = 2  # 최소 2행
        for row in range(3, sheet.max_row + 1):
            if sheet.cell(row=row, column=3).value:  # 3번째 열에 내용이 있는 경우
                max_content_row = row
        
        # 3번째 열의 마지막 행에서 계획 번호 계산
        max_plan_number_in_sheet = max_content_row - 2 if max_content_row > 2 else 0
    
    # 2번째 열의 기존 번호 확인
    current_max_in_col2 = 0
    for row in range(3, sheet.max_row + 1):  # 3행부터 확인
        cell_value = sheet.cell(row=row, column=2).value
        if isinstance(cell_value, int):
            current_max_in_col2 = max(current_max_in_col2, cell_value)
    
    # 부족한 번호들 추가
    if max_plan_number_in_sheet > current_max_in_col2:
        for plan_num in range(current_max_in_col2 + 1, max_plan_number_in_sheet + 1):
            # 2번째 열에 번호 추가
            plan_row = plan_num + 2  # 계획 번호에 해당하는 행
            
            # 2번째 열에 번호 추가 (오른쪽 정렬)
            cell = sheet.cell(row=plan_row, column=2, value=plan_num)
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border
            
            # 해당 행의 다른 열들에도 테두리 적용
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=plan_row, column=col_idx)
                if not cell.border or cell.border == Border():  # 기존 테두리가 없는 경우만
                    cell.border = thin_border

print("모든 시트의 번호 및 테두리 업데이트 완료")
            
for name_idx, name in enumerate(names):
    col = name_idx + 3  # 열 인덱스 (3열부터 시작)
    
    if name not in name_sheets:
        continue
    
    person_sheet = name_sheets[name]
    completed_plans = []
    
    # 해당 사람의 열에서 'O' 표시된 계획 번호 찾기
    for row in range(3, plan_sheet.max_row + 1):  # 3행부터 시트의 마지막 행까지
        cell_value = plan_sheet.cell(row=row, column=col).value
        if cell_value == 'O':
            plan_number = row - 2  # 행 번호에서 계획 번호 계산
            completed_plans.append(plan_number)
    
    # 완료된 계획들에 취소선 적용
    for plan_number in completed_plans:
        plan_row = plan_number + 2  # 계획 번호에서 행 번호 계산
        plan_cell_col3 = person_sheet.cell(row=plan_row, column=3)  # 3번째 열 (계획 내용)
        plan_cell_col2 = person_sheet.cell(row=plan_row, column=2)  # 2번째 열
        
        # 3번째 열 셀에 내용이 있는 경우에만 취소선 적용
        if plan_cell_col3.value:
            # 3번째 열에 취소선 적용
            current_font_col3 = plan_cell_col3.font
            new_font_col3 = Font(
                name=current_font_col3.name,
                size=current_font_col3.size,
                bold=current_font_col3.bold,
                italic=current_font_col3.italic,
                color=current_font_col3.color,
                strike=True  # 취소선 적용
            )
            plan_cell_col3.font = new_font_col3
            
            # 2번째 열에도 취소선 적용 (내용이 있는 경우)
            if plan_cell_col2.value:
                current_font_col2 = plan_cell_col2.font
                new_font_col2 = Font(
                    name=current_font_col2.name,
                    size=current_font_col2.size,
                    bold=current_font_col2.bold,
                    italic=current_font_col2.italic,
                    color=current_font_col2.color,
                    strike=True  # 취소선 적용
                )
                plan_cell_col2.font = new_font_col2
            
            print(f"{name} - {plan_number}번째 계획에 취소선 적용: {plan_cell_col3.value}")

# 저장
wb.save(output_path)
print(f"✅ 완료! 저장된 엑셀 파일: {output_path}")